import discord
from discord.ext import commands
from discord import Embed
import openpyxl
import os
from dotenv import load_dotenv

# MEMBER LIST
workbook = openpyxl.load_workbook('files/member_list.xlsx')

# GENERAL INFORMATION CHANNELS
important_information_channel = int(os.getenv("IMPORTANT_INFORMATION_CHANNEL"))

class positions(commands.Cog):
    def __init__(self, bot):
        self.bot = bot

        async def on_ready():
            print(f'{bot.user} is now running!')

        """ POSITIONS COMMANDS """

        @commands.command(name='positions_exec_board')
        async def positions_exec_board(self, ctx):
            sheet = workbook.active
                
            channel = bot.get_channel(important_information_channel)
                
            branch = f"-----------------------------------------------------------\n**EXECUTIVE BOARD**\n-----------------------------------------------------------"
            target_sheet_name = "executive_board"

            if target_sheet_name in workbook.sheetnames:
                sheet = workbook[target_sheet_name]
                await channel.send(branch)

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    formatted_text = f"**{row[0]}:** {row[1]}"
                    await channel.send(formatted_text)

        @commands.command(name='positions_exec_cabinet')
        async def positions_exec_cabinet(self, ctx):
            sheet = workbook.active
                
            channel = bot.get_channel(important_information_channel)
                
            branch = f"-----------------------------------------------------------\n**EXECUTIVE CABINET**\n-----------------------------------------------------------"
            target_sheet_name = "executive_cabinet"

            if target_sheet_name in workbook.sheetnames:
                sheet = workbook[target_sheet_name]
                await channel.send(branch)

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    formatted_text = f"**{row[0]}:** {row[1]}"
                    await channel.send(formatted_text)

        @commands.command(name='positions_legislative')
        async def positions_legislative(self, ctx):
            sheet = workbook.active

            channel = bot.get_channel(important_information_channel)

            branch = f"-----------------------------------------------------------\n**LEGISLATIVE**\n-----------------------------------------------------------"
            target_sheet_name = "legislative"

            if target_sheet_name in workbook.sheetnames:
                sheet = workbook[target_sheet_name]
                await channel.send(branch)

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    formatted_text = f"**{row[0]}:** {row[1]}"
                    await channel.send(formatted_text)

        @commands.command(name='positions_judicial')
        async def positions_judicial(self, ctx):
            sheet = workbook.active
                
            channel = bot.get_channel(important_information_channel)
                
            branch = f"-----------------------------------------------------------\n**JUDICIAL**\n-----------------------------------------------------------"
            target_sheet_name = "judicial"

            if target_sheet_name in workbook.sheetnames:
                sheet = workbook[target_sheet_name]
                await channel.send(branch)

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    formatted_text = f"**{row[0]}:** {row[1]}"
                    await channel.send(formatted_text)

async def setup(bot):
    await bot.add_cog(positions(bot))